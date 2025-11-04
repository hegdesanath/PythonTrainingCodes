
import openpyxl as xl
import pathlib as pl
import os
from openpyxl.chart import LineChart, Reference, BarChart

os.chdir("C:\\Users\\sanat\\PycharmProjects\\PythonProject\\Datasets\\MobilePriceClassificationDataset")
for file in os.listdir():
    print(file)

def create_excel(fName,sname):
    wb=xl.load_workbook(fName)
    sheet=wb[sname]
    maxCol=sheet.max_column
    maxRow=sheet.max_row
    for col in range(1,maxCol+1):
        if sheet.cell(row=1,column=col).value=="mobile_wt":
            mobile_wt=col
        if sheet.cell(row=1,column=col).value=="sc_h":
            sc_h=col
    #print(f"Mobile Wt Column: {mobile_wt}")
    #print(f"sc_h Column: {sc_h}")
    sheet.cell(1,maxCol+1).value="mobile_wt*sc_h"
    for row in range(2,maxRow+1):
        sheet.cell(row=row,column=maxCol+1).value=((sheet.cell(row=row, column=mobile_wt).value) *
                                                   (sheet.cell(row=row,column=sc_h).value))

    values=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=sheet.max_column,max_col=sheet.max_column)
    chart=LineChart()
    chart.add_data(values)
    sheet.add_chart(chart,"W3")
    #sheet.add_chart(BarChart(values=values, title="Line Chart", Legend=sheet.cell(row=1,column=sheet.max_column).value),"W3")
    wb.save("New"+fName)

create_excel("test.xlsx","test")
create_excel("train.xlsx","train")